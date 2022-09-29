import pyautogui
import pytesseract
import numpy as np
import cv2 as cv
import time
import xlsxwriter
from openpyxl import load_workbook

#from PIL import Image

from general import General
gl = General




class Scraper:
    
    def weeklylootplayers(row_count):
        
        book = load_workbook('Scraper.xlsx')
        sheet = book['Weekly Loot Rankings']
        
        #get into position
        gl.click_button('Rankings')
        x, y = pyautogui.locateCenterOnScreen('images/scraper/Weekly Rankings Next.png', grayscale=False)
        pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
        pyautogui.click()
        xt, yt = pyautogui.locateCenterOnScreen('images/scraper/Top.png', grayscale=False)
        pyautogui.moveTo(xt, yt, .2, pyautogui.easeInQuad)
        pyautogui.click()
        time.sleep(2)
        
        #gather data
        row = 1
        page = 1
        pages = row_count/10
        while page < pages+1:
        
            while row < 11:
                #per page behavior
                name = pyautogui.screenshot('Images/scraper/temp/leaderboard%sname.png' % (row+((page-1)*10)), region=(750, 440+(27*(row-1)), 147, 17))
                alliance = pyautogui.screenshot('Images/scraper/temp/leaderboard%aalliance.png' % (row+((page-1)*10)), region=(900, 440+(27*(row-1)), 144, 17))                          
                loot = pyautogui.screenshot('Images/scraper/temp/leaderboard%sloot.png' % (row+((page-1)*10)), region=(1158, 440+(27*(row-1)), 109, 17))
                
                #preprocessing
                nameimg = name.convert('L')
                nametext = pytesseract.image_to_string(nameimg)
                
                allianceimg = alliance.convert('L')
                alliancetext = pytesseract.image_to_string(allianceimg)     
                          
                lootimg = loot.convert('L')
                loottext = pytesseract.image_to_string(lootimg)
                
                #write data to excel
                sheet.cell(row=3+(row+((page-1)*10)), column=3).value = nametext
                sheet.cell(row=3+(row+((page-1)*10)), column=4).value = alliancetext
                sheet.cell(row=3+(row+((page-1)*10)), column=5).value = loottext
                
                print(nametext,alliancetext,loottext)
                row += 1
                
            x, y = pyautogui.locateCenterOnScreen('images/scraper/Down.png', grayscale=False)
            pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
            pyautogui.click()
            pyautogui.moveTo(xt, yt, .2, pyautogui.easeInQuad)
            time.sleep(.5)
            page += 1
            row = 1
            
           
        print('Processed %s records' % (row_count))
        book.save('Scraper.xlsx')
        
    def alliancelootranking():
        
        book = load_workbook('Scraper.xlsx')
        sheet = book['Alliance Loot Rankings']
        
        
        #gl.click_button('Alliance')
        #x, y = pyautogui.locateCenterOnScreen('images/scraper/Alliance Members List.png', grayscale=False)
        #pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
        #pyautogui.click()
        
        members = pyautogui.screenshot('Images/scraper/temp/members.png', region=(950, 374, 24, 15)) #953, 374, 23, 15
        membersimg = members.convert('L')
        memberstext = pytesseract.image_to_string(membersimg)
        print('%s members' % (memberstext))
        
        row = 1
        #rows = int(memberstext) - 9
        rows = int(memberstext)
        while row < 10:
                name = pyautogui.screenshot('Images/scraper/temp/alliance%sname.png' % (row), region=(752, 454+(30*(row-1)), 137, 15))
                level = pyautogui.screenshot('Images/scraper/temp/alliance%alevel.png' % (row), region=(912, 455+(30*(row-1)), 24, 15)) 
                
                nameimg = name.convert('L')
                nametext = pytesseract.image_to_string(nameimg)
                
                levelimg = level.convert('L')
                leveltext = pytesseract.image_to_string(levelimg)
                
                sheet.cell(row=3+(row), column=3).value = nametext
                sheet.cell(row=3+(row), column=4).value = leveltext
                
                print(nametext,leveltext)
                
                row += 1
        
        while row < rows:
            
            x, y = pyautogui.locateCenterOnScreen('images/scraper/Down.png', grayscale=False, confidence=.95)
            pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
            pyautogui.click()
            name = pyautogui.screenshot('Images/scraper/temp/alliance%sname.png' % (row), region=(752, 694, 137, 15))
            level = pyautogui.screenshot('Images/scraper/temp/alliance%alevel.png' % (row), region=(912, 694, 24, 15)) 
            
            nameimg = name.convert('L')
            nametext = pytesseract.image_to_string(nameimg)
                
            levelimg = level.convert('L')
            leveltext = pytesseract.image_to_string(levelimg)
            
            sheet.cell(row=3+(row), column=3).value = nametext
            sheet.cell(row=3+(row), column=4).value = leveltext
            
            print(nametext,leveltext)
            
            row +=1
            
        print('Processed %s records' % (memberstext))
        book.save('Scraper.xlsx')
        
    def weeklylootplayer(player):
        gl.click_button('Rankings')
        x, y = pyautogui.locateCenterOnScreen('images/scraper/Weekly Rankings Next.png', grayscale=False)
        pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
        pyautogui.click()
        x, y = pyautogui.locateCenterOnScreen('images/scraper/Ranking Search.png', grayscale=False)
        pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
        pyautogui.click()
        pyautogui.typewrite(player, .2)
        x, y = pyautogui.locateCenterOnScreen('images/scraper/Search Button.png', grayscale=False)
        pyautogui.moveTo(x, y, .2, pyautogui.easeInQuad)
        pyautogui.click()
        time.sleep(1)
        name = pyautogui.screenshot('Images/scraper/temp/searchname.png', region=(750, 549, 147, 17))
        alliance = pyautogui.screenshot('Images/scraper/temp/searchalliance.png', region=(900, 549, 144, 17))
        loot = pyautogui.screenshot('Images/scraper/temp/searchloot.png', region=(1158, 549, 109, 17))
        
        print('Processed Records')